#!/usr/bin/env python3
"""
Data pipeline for the VicCamAlertsAuto app.

Downloads the latest "Approved mobile digital road safety camera locations"
spreadsheet from vic.gov.au, geocodes each LOCATION/SUBURB pair (the source
data has no coordinates, only road name + suburb), and writes
app/src/main/assets/cameras.json for the Android app to bundle, plus
data/cameras.json which is what the daily GitHub Actions job publishes for
the app to fetch at runtime (see .github/workflows/update-cameras.yml).

Run this daily via the scheduled workflow, or manually any time. vic.gov.au
only republishes the source file monthly, so most daily runs are a no-op
after the initial pass; geocoding results are cached in geocode_cache.json
keyed by "location|suburb", so re-runs only geocode newly-added or changed
rows -- it does not re-hit the geocoder for roads already resolved.

Uses OpenStreetMap Nominatim (no API key required) at its documented rate
limit of 1 request/second, with a descriptive User-Agent identifying this as
a personal, non-commercial project, per Nominatim's usage policy.
"""
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

TOOLS_DIR = Path(__file__).parent
RAW_XLSX = TOOLS_DIR / "cameras_raw.xlsx"
CACHE_PATH = TOOLS_DIR / "geocode_cache.json"
ASSET_OUTPUT_PATH = TOOLS_DIR.parent / "app" / "src" / "main" / "assets" / "cameras.json"
PUBLISHED_OUTPUT_PATH = TOOLS_DIR.parent / "data" / "cameras.json"

SOURCE_PAGE_URL = "https://www.vic.gov.au/approved-mobile-camera-locations"
USER_AGENT = "VicCamAlertsAuto/1.0 (personal Android Auto project; contact bmsomething98@gmail.com)"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
REQUEST_DELAY_SECONDS = 1.1


def download_latest_spreadsheet() -> None:
    """Scrape the vic.gov.au page for the current month's .xlsx link and download it.

    The filename embeds the month/year (e.g. Mobile-Camera-Locations-July-2026.xlsx)
    and changes each time vic.gov.au republishes, so it can't be hardcoded.
    """
    req = urllib.request.Request(SOURCE_PAGE_URL, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=30) as resp:
        html = resp.read().decode("utf-8", errors="replace")

    match = re.search(r'href="([^"]+\.xlsx)"', html)
    if not match:
        print(f"Could not find an .xlsx link on {SOURCE_PAGE_URL}", file=sys.stderr)
        sys.exit(1)

    file_url = urllib.parse.urljoin(SOURCE_PAGE_URL, match.group(1))
    print(f"Downloading {file_url}")
    file_req = urllib.request.Request(file_url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(file_req, timeout=60) as resp:
        RAW_XLSX.write_bytes(resp.read())


def load_cache() -> dict:
    if CACHE_PATH.exists():
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    return {}


def save_cache(cache: dict) -> None:
    CACHE_PATH.write_text(json.dumps(cache, indent=2), encoding="utf-8")


def geocode(query: str) -> dict | None:
    params = {
        "q": query,
        "format": "jsonv2",
        "limit": 1,
        "countrycodes": "au",
    }
    url = f"{NOMINATIM_URL}?{urllib.parse.urlencode(params)}"
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            results = json.loads(resp.read().decode("utf-8"))
    except Exception as exc:
        print(f"  ! geocode error for {query!r}: {exc}", file=sys.stderr)
        return None
    if not results:
        return None
    r = results[0]
    return {"lat": float(r["lat"]), "lon": float(r["lon"])}


def geocode_row(location: str, suburb: str, cache: dict) -> dict:
    key = f"{location}|{suburb}"
    if key in cache:
        return cache[key]

    # Try street + suburb first (most precise), then fall back to suburb only.
    result = geocode(f"{location}, {suburb}, Victoria, Australia")
    precision = "street"
    if result is None:
        time.sleep(REQUEST_DELAY_SECONDS)
        result = geocode(f"{suburb}, Victoria, Australia")
        precision = "suburb"

    time.sleep(REQUEST_DELAY_SECONDS)

    if result is None:
        entry = {"lat": None, "lon": None, "precision": "unresolved"}
    else:
        entry = {"lat": result["lat"], "lon": result["lon"], "precision": precision}

    cache[key] = entry
    return entry


def main() -> None:
    if "--use-existing" in sys.argv and RAW_XLSX.exists():
        print(f"Using existing {RAW_XLSX} (--use-existing passed)")
    else:
        download_latest_spreadsheet()

    wb = openpyxl.load_workbook(RAW_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    rows = [r for r in rows if r and r[0]]

    cache = load_cache()
    cameras = []
    new_geocodes = 0

    for i, (location, suburb, reason_code, audit_date) in enumerate(rows, start=1):
        key = f"{location}|{suburb}"
        was_cached = key in cache
        entry = geocode_row(location, suburb, cache)
        if not was_cached:
            new_geocodes += 1
            if new_geocodes % 25 == 0:
                save_cache(cache)  # checkpoint periodically in case of interruption

        if entry["lat"] is None:
            continue  # skip rows we truly could not resolve

        cameras.append(
            {
                "location": location,
                "suburb": suburb,
                "lat": entry["lat"],
                "lon": entry["lon"],
                "precision": entry["precision"],
                "reasonCode": reason_code,
                "auditDate": audit_date,
            }
        )

        if i % 100 == 0 or i == len(rows):
            print(f"[{i}/{len(rows)}] processed ({new_geocodes} newly geocoded this run)")

    save_cache(cache)

    payload = json.dumps(cameras, indent=None, separators=(",", ":"))
    for output_path in (ASSET_OUTPUT_PATH, PUBLISHED_OUTPUT_PATH):
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(payload, encoding="utf-8")

    unresolved = len(rows) - len(cameras)
    print(
        f"Wrote {len(cameras)} cameras to {ASSET_OUTPUT_PATH} and {PUBLISHED_OUTPUT_PATH} "
        f"({unresolved} unresolved/skipped)."
    )


if __name__ == "__main__":
    main()
